import logging
import os
from pathlib import Path
from typing import Any, List

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.ireader import IReader

logger = logging.getLogger(__name__) # 放置在模块顶部


class Xlsx(IReader):
	"""xlsx解析封装
	默认打开第一个sheet
	无错误处理,可能会抛出异常

	self.__workbook和self.__workbookW的区别是打开表格时是否存在参数data_only
	data_only=True则不会读取公式,直接会读取公式的计算结果,直接使用这份数据保存会导致公式丢失
	data_only=False则会读取公式,取数据时,取到的是公式,不会计算单元格的值

	- 修改表格数据, `eg_xlsx_write()`
	- 读取表格数据, `eg_xlsx_read()`
	"""

	def __init__(self, filename: str, ignore_fomula: bool = False, init_write: bool = False) -> None:
		"""xlsx解析封装

		Args:
			filename (str): 文件名
			ignore_fomula (bool): 是否忽略公式, 忽略公式后, 原表格中所有的公式会被替换成计算后的值, 保存excel后原表格的公式会丢失替换为实际值
			init_write (bool): 是否直接初始化写表格数据, 在有预加载表格的情况下, 启用这个功能可以提升写表格效率
		"""
		self.__filename = filename
		self.__ignore_fomula = ignore_fomula
		self.__workbookW: Workbook = None # type: ignore # 写数据专用表
		self.__sheetW: Worksheet = None # type: ignore # 写数据专用sheet

		# 被修改的数据记录, 恢复数据用, 只能记录修改数据, 删除行不可用 <row_index, <col_index, value>>
		self.__changed_datas: dict[int, dict[int, Any]] = {}

		self.__workbook: Workbook = None # type: ignore
		try:
			self.__workbook = load_workbook(self.__filename, data_only=True)
			self.sheet(0)
			if init_write:
				self.__try_init_xlsx_write()
		except (FileNotFoundError, InvalidFileException) as e:
			logger.error(f"加载Xlsx报错(文件没找到, 无效文件), 文件名: {filename}, 错误: {e}")
		except Exception as e:
			logger.error(f"加载Xlsx报错(其他), 文件名: {filename}, 错误: {e}")

	def is_open(self) -> bool:
		return self.__workbook is not None

	def sheet(self, index: int) -> None:
		"""根据索引打开sheet

		Args:
			index (int): 打开的sheet索引,从0开始
		"""
		self.__sheet_index = index
		self.__sheet = self.__workbook.worksheets[index]

		if self.__workbookW is not None:
			self.__sheetW = self.__workbookW.worksheets[index]

	def get_max_row(self) -> int:
		"""获取表格最大行数

		Returns:
			int: 表格最大行数
		"""
		return self.__sheet.max_row

	def get_max_column(self) -> int:
		"""获取表格最大列数

		Returns:
			int: 表格最大列数
		"""
		return self.__sheet.max_column

	def append_row(self, new_row: List[Any]) -> None:
		"""追加一行数据到表尾

		Args:
			new_row (List[Any]): 追加行数据,从头开始按照顺序插入
		"""
		max_row = self.get_max_row()
		self.insert_row(max_row + 1, new_row)

	def insert_row(self, row_index: int, new_row: List[Any]) -> None:
		"""插入一行数据

		Args:
			row_index (int): 插入行位置,从1开始,就是excel最左侧的行数
			new_row (List[Any]): 插入行数据,从头开始按照顺序插入
		"""
		self.__try_init_xlsx_write()

		self.__sheet.insert_rows(row_index)
		if self.__sheetW:
			self.__sheetW.insert_rows(row_index)

		for col_index, value in enumerate(new_row, start=1):
			self.set_value(row_index, col_index, value)

		self.copy_row_styles(row_index - 1, row_index)

	def delete_row(self, row_index: int) -> None:
		"""删除一行

		Args:
			row_index (int): 删除行位置,从1开始
		"""
		# raise Exception("delete_role function denied") # 当前禁止删除行

		self.__try_init_xlsx_write()

		self.__sheet.delete_rows(row_index)
		if self.__sheetW:
			self.__sheetW.delete_rows(row_index)

	def clear_row(self, row_index: int) -> None:
		"""清空一行数据, 保留空行

		Args:
			row_index (int): 清空行位置,从1开始
		"""

		col_num = self.get_max_column()
		for col_index in range(1, col_num + 1):
			self.set_value(row_index, col_index, "")

	def set_row(self, row_index: int, new_row: List[Any]) -> None:
		"""设置一行数据

		Args:
			row_index (int): 设置行位置,从1开始
			new_row (List[Any]): 填充行数据,从头开始按照顺序插入
		"""

		l = len(new_row)
		col_num = self.get_max_column()
		m = min(col_num, l)
		for col_index in range(1, m + 1):
			self.set_value(row_index, col_index, new_row[col_index - 1])

	def get_value(self, row_index: int, col_index: int) -> Any:
		"""获取单元格值

		Args:
			row_index (int): 行数,从1开始
			col_index (int): 列数,从1开始

		Returns:
			Any: 单元格值
		"""
		return self.__sheet.cell(row=row_index, column=col_index).value

	def set_value(self, row_index: int, col_index: int, value: Any) -> None:
		"""设置单元格值

		Args:
			row_index (int): 行数,从1开始
			col_index (int): 列数,从1开始
			value (Any): 值
		"""
		self.__try_init_xlsx_write()
		self.__try_record_change_data(row_index, col_index)

		self.__sheet.cell(row=row_index, column=col_index, value=value)
		if self.__sheetW:
			self.__sheetW.cell(row=row_index, column=col_index, value=value)

	def copy_row_styles(self, src_row_index: int, tar_row_index: int) -> None:
		"""复制目标行的样式"""

		def copy_styles(sheet: Worksheet | None):
			if sheet is None:
				return
			src_row = list(sheet[src_row_index])
			tar_row = list(sheet[tar_row_index])
			for source_cell, target_cell in zip(src_row, tar_row):
				if source_cell.has_style:
					try:
						target_cell.font = source_cell.font.copy()
						target_cell.border = source_cell.border.copy()
						target_cell.fill = source_cell.fill.copy()
						target_cell.number_format = source_cell.number_format
						target_cell.protection = source_cell.protection.copy()
						target_cell.alignment = source_cell.alignment.copy()
					except Exception as e:
						logger.error(f"Error copying style from {source_cell} to {target_cell}: {e}")
						# raise Exception(f"Error copying style from {source_cell} to {target_cell}: {e}")

		if src_row_index <= 0:
			return

		copy_styles(self.__sheet)
		copy_styles(self.__sheetW)

	def save(self) -> None:
		"""保存"""

		self.save_as(self.__filename)

	def save_as(self, filename: str) -> None:
		"""保存为文件"""

		if len(self.__changed_datas) == 0: # 没有数据变动则不需要保存
			return

		if self.__workbookW is None:
			if self.__ignore_fomula:
				self.__workbook.save(filename=filename)
				return
			else:
				return # 打开一个表, 然后直接写表就会到这里
				# raise AttributeError("self.__workbookW is None")

		self.__workbookW.save(filename=filename)

	def restore(self) -> None:
		"""恢复变更数据"""

		if len(self.__changed_datas) == 0:
			return

		for row_index, datas in self.__changed_datas.items():
			for col_index, value in datas.items():
				self.set_value(row_index, col_index, value) # 这里实际不会引起重复调用, 不会有问题

		self.drop_changed_data()

	def drop_changed_data(self) -> None:
		"""丢弃变更前原始数据"""

		self.__changed_datas.clear()

	def debug(self) -> None:
		print("xlsx: ", self.__get_fullpath())

		mr = self.get_max_row()
		mc = self.get_max_column()

		for r in range(1, mr + 1):
			info = [str(r)]
			for c in range(1, mc + 1):
				v = self.get_value(r, c)
				info.append(str(v) if v else "_empty_")
			print(",\t".join(info))

	def __get_fullpath(self) -> str:
		if os.path.isabs(self.__filename):
			return self.__filename
		return str((Path(os.getcwd()) / self.__filename).resolve())

	def __try_init_xlsx_write(self):
		if self.__ignore_fomula:
			return
		if self.__workbookW is not None:
			return
		self.__workbookW = load_workbook(self.__filename)

		self.sheet(self.__sheet_index)

	def __try_record_change_data(self, row_index: int, col_index: int) -> None:
		"""尝试记录变更的数据, 只会记录第一次变更的数据"""

		if row_index not in self.__changed_datas:
			self.__changed_datas[row_index] = {}
		if col_index in self.__changed_datas[row_index]: # 已经记录过了
			return

		self.__changed_datas[row_index][col_index] = self.get_value(row_index, col_index)
